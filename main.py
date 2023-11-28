import openpyxl
inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file['Sheet1']
product_per_supplier = {}
total_value_per_supplier = {}
product_with_inventory_less_than_ten = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)


# CALCULATION NUMBER OF PRODUCTS PER SUPPLIER

    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

# Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # inventory in product_with_inventory_less_than_ten:
    if inventory < 10:
        product_with_inventory_less_than_ten[product_num] = inventory

# add value for total inventory price-value
    inventory_price.value = inventory * price
# inventory_file.save("inventory_with_total value.xlsx")
print(product_per_supplier)
print(total_value_per_supplier)
print(product_with_inventory_less_than_ten)

